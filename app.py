import os
import re
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
from sklearn.cluster import AgglomerativeClustering
from fpdf import FPDF
from collections import defaultdict
from collections import Counter

# ========== CONFIG ==========
SUBMISSIONS_DIR = "submissions"
REPORT_FILE = "ai_detection_report.csv"

# similarity thresholds
SIMILARITY_THRESHOLD = 0.9            # general text/formula duplicate threshold
RELATIVE_SIMILARITY_THRESHOLD = 0.8   # threshold for highlighting rare-but-shared formulas

# suspicious score weights
WEIGHT_FORMULA_DUP = 3
WEIGHT_TEXT_DUP = 2
WEIGHT_AI_METADATA = 5
# ============================


def compute_anomaly_scores(series):
    """
    Compute a rarity/anomaly score for each entry in a pandas Series.
    Common values -> score near 0
    Rare values -> score near 1
    """
    counts = Counter(series)
    total = len(series)
    return series.map(lambda val: 1 - counts[val] / total)

def extract_excel_data(filepath):
    """Extract metadata, text, and formulas from an Excel file."""
    data = {
        "filename": os.path.basename(filepath),
        "creator": None,
        "lastModifiedBy": None,
        "created": None,
        "modified": None,
        "text_content": "",
        "formula_content": "",
        "num_sheets": 0
    }

    try:
        wb = load_workbook(filepath, data_only=False)
        props = wb.properties
        data.update({
            "creator": props.creator,
            "lastModifiedBy": props.lastModifiedBy,
            "created": props.created,
            "modified": props.modified,
            "num_sheets": len(wb.sheetnames)
        })

        text_cells, formula_cells = [], []

        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(values_only=False):
                for cell in row:
                    if cell.data_type == "s" and cell.value:
                        text_cells.append(str(cell.value))
                    elif cell.data_type == "f" and cell.value:
                        formula_cells.append(str(cell.value))

        data["text_content"] = " ".join(text_cells)
        data["formula_content"] = " ".join(formula_cells)

    except Exception as e:
        data["error"] = str(e)

    return data

def compute_similarity(contents, token_pattern=r"(?u)\b\w+\b"):
    """
    Compute cosine similarity for a list of strings safely.
    Default token_pattern matches words.
    """
    # Replace NaN with empty string
    safe_contents = [str(c) if c else "" for c in contents]

    vectorizer = TfidfVectorizer(token_pattern=token_pattern, stop_words="english")
    tfidf = vectorizer.fit_transform(safe_contents)
    return cosine_similarity(tfidf)

def find_duplicates(sim_matrix, filenames, threshold=SIMILARITY_THRESHOLD):
    duplicates = []
    n = len(filenames)
    for i in range(n):
        for j in range(i + 1, n):
            if sim_matrix[i, j] >= threshold:
                duplicates.append((filenames[i], filenames[j], round(sim_matrix[i, j], 3)))
    return duplicates

def detect_metadata_anomalies(df):
    anomalies = []

    # Identical creator/lastModifiedBy
    grouped = df.groupby(["creator", "lastModifiedBy"]).size()
    for idx, count in grouped.items():
        if count > 1:
            anomalies.append(f"{count} files share identical metadata: {idx}")

    # Keywords indicating AI usage
    for _, row in df.iterrows():
        if row["creator"] and any(k in str(row["creator"]).lower() for k in ["copilot", "chatgpt", "claude", "openai"]):
            anomalies.append(f"{row['filename']} references AI in metadata ({row['creator']})")

    return anomalies

def cluster_submissions(sim_matrix, filenames):
    clusters = defaultdict(list)
    if len(filenames) < 2:
        clusters[0] = filenames
        return clusters

    dist_matrix = 1 - sim_matrix
    dist_matrix = np.nan_to_num(dist_matrix)

    clustering = AgglomerativeClustering(
        n_clusters=None,
        distance_threshold=1 - SIMILARITY_THRESHOLD,
        metric='precomputed',
        linkage='average'
    )
    labels = clustering.fit_predict(dist_matrix)
    for file, label in zip(filenames, labels):
        clusters[label].append(file)
    return clusters

def find_relative_duplicates(sim_matrix, filenames, min_shared=2, max_shared_ratio=0.1, threshold=RELATIVE_SIMILARITY_THRESHOLD):
    """
    Identify clusters of submissions that share a formula/text more than `min_shared` 
    but less than `max_shared_ratio` of the total submissions.
    """
    n = len(filenames)
    clusters = []
    visited = set()

    for i in range(n):
        if i in visited:
            continue
        similar = [i] + [j for j in range(n) if j != i and sim_matrix[i, j] > threshold]
        if min_shared <= len(similar) <= max(1, int(max_shared_ratio * n)):
            clusters.append([filenames[k] for k in similar])
            visited.update(similar)
    return clusters


def analyze_excel_folder(submissions_dir):
    """Analyze all Excel submissions in a folder."""
    records = []
    for filename in os.listdir(submissions_dir):
        if filename.endswith(".xlsx"):
            path = os.path.join(submissions_dir, filename)
            record = extract_excel_data(path)
            record["student_name"] = record.get("lastModifiedBy") or record.get("creator") or "Unknown"
            records.append(record)

    df = pd.DataFrame(records)
    if len(df) == 0:
        return {
            "report_text": "No valid Excel files found.",
            "df": df,
            "text_dups": [],
            "formula_dups": [],
            "formula_dups_relative": [],
            "metadata_flags": [],
            "clusters": {}
        }

    # --- Compute similarities ---
    text_sim = compute_similarity(df["text_content"].fillna(""))
    formula_sim = compute_similarity(
        df["formula_content"].fillna(""), 
        token_pattern=r"[\w\+\-\*/\(\)]+"
    )

    formula_dups = find_duplicates(formula_sim, df["filename"])

    # --- Detect duplicates ---
    text_dups = find_duplicates(text_sim, df["filename"])
    formula_dups_relative = find_relative_duplicates(
        formula_sim,
        df["filename"],
        threshold=RELATIVE_SIMILARITY_THRESHOLD
    )

    # --- Compute relative formula similarity (anomaly-based) ---
    from collections import Counter
    import numpy as np

    def compute_anomaly_scores(series):
        counts = Counter(series)
        total = len(series)
        return series.map(lambda val: 1 - counts[val] / total)

    anomaly_scores = compute_anomaly_scores(df["formula_content"].fillna(""))
    n = len(df)
    relative_similarity = np.zeros((n, n))
    for i in range(n):
        for j in range(i+1, n):
            avg_anomaly = (anomaly_scores.iloc[i] + anomaly_scores.iloc[j]) / 2
            relative_similarity[i, j] = formula_sim[i, j] * avg_anomaly
            relative_similarity[j, i] = relative_similarity[i, j]

    suspicious_threshold = 0.7
    formula_dups_relative = find_duplicates(relative_similarity, df["filename"], threshold=suspicious_threshold)

    # --- Metadata anomalies and clustering ---
    metadata_flags = detect_metadata_anomalies(df)
    clusters = cluster_submissions(formula_sim, df["filename"])

    # --- Suspicious score ---
    suspicious_scores = []
    for _, row in df.iterrows():
        score = 0
        if any(row["filename"] in dup for dup in formula_dups):
            score += 3
        if any(row["filename"] in dup for dup in text_dups):
            score += 2
        if any(row["filename"] in dup for dup in formula_dups_relative):
            score += 4  # extra weight for rare-but-shared formulas
        if row.get("creator") and any(k in str(row["creator"]).lower() for k in ["copilot", "chatgpt", "claude", "openai"]):
            score += 5
        suspicious_scores.append(score)
    df["suspicious_score"] = suspicious_scores
    df = df.sort_values("suspicious_score", ascending=False)

    # --- Save CSV ---
    df.to_csv(os.path.join(submissions_dir, REPORT_FILE), index=False)

    # --- Build summary text ---
    summary = []
    summary.append(f"✅ Total submissions: {len(df)}")
    summary.append(f"⚠️ Text duplicates: {len(text_dups)} | Formula duplicates: {len(formula_dups)} | Relative formula duplicates: {len(formula_dups_relative)}")
    summary.append(f"📁 Clusters: {sum(len(v) > 1 for v in clusters.values())}")
    summary.append(f"🔎 Metadata anomalies: {len(metadata_flags)}\n")
    summary.append("Top suspicious submissions:")
    for _, r in df.head(5).iterrows():
        summary.append(f" - {r['student_name']} ({r['filename']}): score {r['suspicious_score']}")
    if metadata_flags:
        summary.append("\nMetadata flags:")
        for f in metadata_flags:
            summary.append(f" - {f}")

    report_text = "\n".join(summary)
    return {
        "report_text": report_text,
        "df": df,
        "text_dups": text_dups,
        "formula_dups": formula_dups,
        "formula_dups_relative": formula_dups_relative,
        "metadata_flags": metadata_flags,
        "clusters": clusters
    }

def strip_non_ascii(text):
    return re.sub(r"[^\x00-\x7F]", "?", text)

def break_long_words(text, max_len=80):
    def split_word(word):
        return "\n".join([word[i:i+max_len] for i in range(0, len(word), max_len)])
    return " ".join(split_word(w) if len(w) > max_len else w for w in text.split())

def safe_text_for_pdf(text):
    """Ensure text is a string and replace problematic characters for PDF."""
    if text is None:
        return ""
    return str(text).replace("\r", "").replace("\t", "    ")

def create_pdf_report(df, text_dups, formula_dups, formula_dups_relative, metadata_flags, clusters, output_path):

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    pdf.set_font("Arial", "B", 16)
    pdf.multi_cell(0, 10, "LBO Model AI Report - PE Methods", align="C")
    pdf.ln(5)

    pdf.set_font("Arial", size=12)
    summary_lines = [
        f"✅ Total submissions: {len(df)}",
        f"⚠️ Text duplicates: {len(text_dups)}",
        f"⚠️ Formula duplicates: {len(formula_dups)}",
        f"⚠️ Relative formula duplicates (rare-but-shared): {len(formula_dups_relative)}",
        f"📁 Clusters: {sum(len(v) > 1 for v in clusters.values())}",
        f"🔎 Metadata anomalies: {len(metadata_flags)}"
    ]

    for line in summary_lines:
        safe_line = safe_text_for_pdf(line)
        for subline in safe_line.split("\n"):
            pdf.multi_cell(0, 6, subline)
        pdf.ln(2)

    if "suspicious_score" in df.columns:
        pdf.add_page()
        pdf.set_font("Arial", "B", 12)
        pdf.multi_cell(0, 8, "Top Suspicious Submissions:")
        pdf.ln(2)

        pdf.set_font("Arial", size=11)
        for _, row in df.head(10).iterrows():
            line = f"{row.get('student_name', 'Unknown')} ({row.get('filename', 'Unknown')}) - Suspicious Score: {row.get('suspicious_score', 'N/A')}"
            for subline in safe_text_for_pdf(line).split("\n"):
                pdf.multi_cell(0, 6, subline)
            pdf.ln(1)

    if formula_dups_relative:
        pdf.add_page()
        pdf.set_font("Arial", "B", 12)
        pdf.multi_cell(0, 8, "Relative Formula Duplicates (rare-but-shared):")
        pdf.ln(2)
        pdf.set_font("Arial", size=11)
        for dup_group in formula_dups_relative:
            group_line = ", ".join(dup_group)
            for subline in safe_text_for_pdf(group_line).split("\n"):
                pdf.multi_cell(0, 6, subline)
            pdf.ln(1)

    if metadata_flags:
        pdf.add_page()
        pdf.set_font("Arial", "B", 12)
        pdf.multi_cell(0, 8, "Metadata Flags:")
        pdf.ln(2)
        pdf.set_font("Arial", size=11)
        for flag in metadata_flags:
            for subline in safe_text_for_pdf(flag).split("\n"):
                pdf.multi_cell(0, 6, subline)
            pdf.ln(1)

    pdf.output(output_path)
    return output_path

# Optional: allow running from terminal
if __name__ == "__main__":
    results = analyze_excel_folder(SUBMISSIONS_DIR)
    print(results["report_text"])